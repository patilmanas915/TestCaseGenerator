import csv
import os
from datetime import datetime
from config import Config
from gemini_client import GeminiClient
import json
import logging

# Using pandas-free implementation for Render compatibility
PANDAS_AVAILABLE = False

logger = logging.getLogger(__name__)

class TestCaseGenerator:
    def __init__(self):
        self.gemini_client = GeminiClient()
        self.all_test_cases = []
        
    def process_frd_document(self, frd_content, progress_callback=None):
        """Process FRD document and generate test cases"""
        
        logger.info("Extracting features from FRD document...")
        if progress_callback:
            progress_callback("Extracting features from FRD document...")
            
        features_data = self.gemini_client.extract_features_from_frd(frd_content)
        
        if not features_data or 'features' not in features_data:
            logger.error("Failed to extract features from FRD")
            return False, "Failed to extract features from FRD document"
            
        features = features_data['features']
        logger.info(f"Found {len(features)} features")
        
        if progress_callback:
            progress_callback(f"Found {len(features)} features. Generating test cases...")
        
        # Generate test cases for each feature
        for idx, feature in enumerate(features, 1):
            # Get feature name properly - this was working correctly
            feature_name = feature.get('feature_name', feature.get('name', f'Feature_{idx}'))
            
            logger.info(f"Generating test cases for Feature {idx}/{len(features)}: {feature_name}")
            
            if progress_callback:
                progress_callback(f"Generating test cases for feature {idx}/{len(features)}: {feature_name}")
            
            test_cases_data = self.gemini_client.generate_test_cases_for_feature(feature)
            
            if test_cases_data and 'test_cases' in test_cases_data:
                # Add feature information to each test case
                for test_case in test_cases_data['test_cases']:
                    test_case['feature_name'] = feature_name
                    test_case['feature_id'] = feature.get('feature_id', feature.get('id', f'F{idx:03d}'))
                    test_case['module'] = feature.get('module', test_case.get('module', 'Unknown'))
                
                self.all_test_cases.extend(test_cases_data['test_cases'])
                logger.info(f"Generated {len(test_cases_data['test_cases'])} test cases for {feature_name}")
            else:
                logger.warning(f"Failed to generate test cases for {feature_name}")
            
            # Add delay to respect API rate limits
            self.gemini_client.rate_limit_delay()
        
        return True, f"Successfully generated {len(self.all_test_cases)} test cases"
    
    def save_to_csv(self, filename=None):
        """Save all generated test cases to Excel file in few-shot format with proper formatting"""
        
        if not self.all_test_cases:
            logger.error("No test cases to save")
            return None, "No test cases to save"
        
        # Prepare data for Excel
        csv_data = []
        
        for idx, test_case in enumerate(self.all_test_cases, 1):
            # Use the formatted test steps from few-shot prompting
            test_steps_formatted = test_case.get('test_steps_formatted', '')
            
            # If not available, convert test steps list to numbered string
            if not test_steps_formatted and test_case.get('test_steps'):
                test_steps_formatted = '\n'.join([
                    f"{i+1}. {step}" 
                    for i, step in enumerate(test_case.get('test_steps', []))
                ])
            
            # Clean and format text fields to avoid Excel formatting issues
            def clean_text(text):
                if not text:
                    return ""
                # Remove extra whitespace and ensure proper line breaks
                cleaned = str(text).strip()
                # Replace multiple spaces with single space
                cleaned = ' '.join(cleaned.split())
                return cleaned
            
            csv_row = {
                'Test_Case_ID': test_case.get('test_case_id', f'TC{idx:03d}'),
                'Test_Case_Name': clean_text(test_case.get('test_case_name', '')),
                'Feature_ID': clean_text(test_case.get('feature_id', '')),
                'Feature_Name': clean_text(test_case.get('feature_name', '')),
                'Module': clean_text(test_case.get('module', '')),
                'Test_Type': clean_text(test_case.get('test_type', '')),
                'Priority': clean_text(test_case.get('priority', '')),
                'Category': clean_text(test_case.get('category', '')),
                'Gap_Coverage': clean_text(test_case.get('gap_coverage', '')),
                'Preconditions': clean_text(test_case.get('preconditions', '')),
                'Steps': test_steps_formatted,  # Changed from Test_Steps_Few_Shot_Format
                'Test_Data': clean_text(test_case.get('test_data', '')),
                'Expected_Result': clean_text(test_case.get('expected_result', '')),
                'FRD_Reference': clean_text(test_case.get('frd_reference', '')),
                'Generated_On': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            csv_data.append(csv_row)
        
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'generated_testcases_{timestamp}.xlsx'
        else:
            # Ensure .xlsx extension
            if not filename.endswith('.xlsx'):
                filename = filename.replace('.csv', '.xlsx')
        
        # Ensure download directory exists
        os.makedirs(Config.DOWNLOAD_FOLDER, exist_ok=True)
        
        # Save to Excel with proper formatting
        output_path = os.path.join(Config.DOWNLOAD_FOLDER, filename)
        
        try:
            # Use openpyxl directly without pandas  
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Test_Cases"
            
            # Add headers
            if csv_data:
                headers = list(csv_data[0].keys())
                ws.append(headers)
                
                # Add data
                for row_data in csv_data:
                    row = [row_data.get(header, '') for header in headers]
                    ws.append(row)
                
                # Auto-adjust column widths with better multi-line content handling
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        if cell.value:
                            # Handle multi-line content by checking line length
                            cell_value = str(cell.value)
                            if '\n' in cell_value:
                                # For multi-line content, use the longest line
                                lines = cell_value.split('\n')
                                max_line_length = max(len(line) for line in lines) if lines else 0
                                cell_length = max_line_length
                            else:
                                cell_length = len(cell_value)
                            
                            if cell_length > max_length:
                                max_length = cell_length
                    
                    # Set column width with better sizing for different content types
                    if max_length < 15:
                        adjusted_width = max_length + 5  # Short content gets more padding
                    elif max_length < 50:
                        adjusted_width = max_length + 3  # Medium content
                    else:
                        adjusted_width = min(max_length + 2, 100)  # Long content, max 100 chars
                    
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Set specific minimum widths for important columns
                column_widths = {
                    'Steps': 80,  # Steps column needs more space
                    'Test_Case_Name': 40,
                    'Expected_Result': 50,
                    'Description': 40,
                    'Preconditions': 30
                }
                
                # Apply minimum widths
                for col_idx, header in enumerate(headers, 1):
                    column_letter = ws.cell(1, col_idx).column_letter
                    if header in column_widths:
                        current_width = ws.column_dimensions[column_letter].width
                        min_width = column_widths[header]
                        ws.column_dimensions[column_letter].width = max(current_width, min_width)
                
                # Auto-adjust row heights for multi-line content
                for row in ws.iter_rows(min_row=2):  # Skip header row
                    max_lines = 1
                    for cell in row:
                        if cell.value and '\n' in str(cell.value):
                            lines = str(cell.value).count('\n') + 1
                            max_lines = max(max_lines, lines)
                        
                        # Enable text wrapping for all cells
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Set row height based on content (15 points per line)
                    ws.row_dimensions[row[0].row].height = max_lines * 15
                
                # Add header formatting
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(output_path)
        
        except Exception as e:
            logger.error(f"Could not create Excel file: {e}")
            return None, f"Could not create Excel file: {e}"
        
        logger.info(f"Successfully saved {len(csv_data)} test cases to {output_path}")
        return output_path, f"Successfully saved {len(csv_data)} test cases in Excel format"
    
    def get_summary_stats(self):
        """Generate summary statistics of generated test cases"""
        
        if not self.all_test_cases:
            return {}
        
        stats = {
            'total_test_cases': len(self.all_test_cases),
            'test_types': {},
            'priorities': {},
            'features': {},
            'modules': {}
        }
        
        # Calculate stats manually
        for test_case in self.all_test_cases:
            test_type = test_case.get('test_type', 'Unknown')
            priority = test_case.get('priority', 'Unknown')
            feature = test_case.get('feature_name', 'Unknown')
            module = test_case.get('module', 'Unknown')
            
            # Count test types
            stats['test_types'][test_type] = stats['test_types'].get(test_type, 0) + 1
            
            # Count priorities
            stats['priorities'][priority] = stats['priorities'].get(priority, 0) + 1
            
            # Count features
            stats['features'][feature] = stats['features'].get(feature, 0) + 1
            
            # Count modules
            stats['modules'][module] = stats['modules'].get(module, 0) + 1
        
        return stats
