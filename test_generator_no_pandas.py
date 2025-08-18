import csv
import json
import os
from datetime import datetime
from config import Config
from gemini_client import GeminiClient
import logging

logger = logging.getLogger(__name__)

class TestCaseGenerator:
    """Pandas-free test case generator for Render deployment"""
    
    def __init__(self):
        self.gemini_client = GeminiClient()
        self.all_test_cases = []
        
    def process_frd_document(self, frd_content, progress_callback=None):
        """Process FRD document and generate test cases"""
        
        logger.info("Extracting features from FRD document...")
        if progress_callback:
            progress_callback("Extracting features from FRD document...")
            
        features_data = self.gemini_client.extract_features_from_frd(frd_content)
        
        if not features_data or 'features' not in features_data:
            logger.error("Failed to extract features from FRD")
            return False, "Failed to extract features from FRD document"
            
        features = features_data['features']
        logger.info(f"Found {len(features)} features")
        
        if progress_callback:
            progress_callback(f"Found {len(features)} features. Generating test cases...")
        
        # Generate test cases for each feature
        all_test_cases = []
        for i, feature in enumerate(features):
            try:
                if progress_callback:
                    progress_callback(f"Generating test cases for feature {i+1}/{len(features)}: {feature.get('name', 'Unknown')}")
                
                test_cases = self.gemini_client.generate_test_cases_for_feature(feature)
                
                if test_cases and 'test_cases' in test_cases:
                    for test_case in test_cases['test_cases']:
                        # Enhanced feature name handling
                        feature_name = feature.get('feature_name', feature.get('name', f'Feature_{i+1}'))
                        feature_id = feature.get('feature_id', feature.get('id', f'F{str(i+1).zfill(3)}'))
                        
                        # Ensure proper field mapping
                        test_case['feature_name'] = feature_name
                        test_case['feature_id'] = feature_id
                        test_case['module'] = feature.get('module', test_case.get('module', 'Unknown'))
                        
                        # Fix field name inconsistencies
                        if 'test_case_id' not in test_case and 'id' in test_case:
                            test_case['test_case_id'] = test_case['id']
                        if 'test_case_name' not in test_case and 'name' in test_case:
                            test_case['test_case_name'] = test_case['name']
                        
                        all_test_cases.append(test_case)
                        
                    logger.info(f"Generated {len(test_cases['test_cases'])} test cases for feature: {feature_name}")
                else:
                    logger.warning(f"No test cases generated for feature: {feature.get('name', 'Unknown')}")
                        
            except Exception as e:
                feature_name = feature.get('feature_name', feature.get('name', 'Unknown'))
                logger.error(f"Error generating test cases for feature {feature_name}: {e}")
                continue
        
        self.all_test_cases = all_test_cases
        logger.info(f"Generated {len(all_test_cases)} total test cases")
        
        if progress_callback:
            progress_callback(f"Generated {len(all_test_cases)} test cases successfully!")
        
        return True, f"Successfully generated {len(all_test_cases)} test cases"
    
    def get_statistics(self):
        """Get statistics about generated test cases"""
        if not self.all_test_cases:
            return {}
        
        stats = {
            'total_test_cases': len(self.all_test_cases),
            'positive_test_cases': 0,
            'negative_test_cases': 0,
            'boundary_test_cases': 0,
            'edge_test_cases': 0,
            'features_covered': len(set([tc.get('feature_name', '') for tc in self.all_test_cases]))
        }
        
        for test_case in self.all_test_cases:
            test_type = test_case.get('type', '').lower()
            if 'positive' in test_type:
                stats['positive_test_cases'] += 1
            elif 'negative' in test_type:
                stats['negative_test_cases'] += 1
            elif 'boundary' in test_type:
                stats['boundary_test_cases'] += 1
            elif 'edge' in test_type:
                stats['edge_test_cases'] += 1
        
        return stats
    
    def save_to_csv(self):
        """Save test cases to CSV format"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"generated_testcases_{timestamp}.csv"
            
            # Get download folder - try config first, then fallbacks
            try:
                download_folder = getattr(Config, 'DOWNLOAD_FOLDER', 'downloads')
                # Convert relative to absolute if needed
                if not os.path.isabs(download_folder):
                    download_folder = os.path.join(os.getcwd(), download_folder)
            except:
                download_folder = '/tmp/downloads'
                
            os.makedirs(download_folder, exist_ok=True)
            filepath = os.path.join(download_folder, filename)
            
            print(f"💾 Saving CSV to: {filepath}")
            
            # Define CSV columns
            fieldnames = [
                'Test_Case_ID', 'Test_Case_Name', 'Feature_ID', 'Feature_Name',
                'Module', 'Test_Type', 'Priority', 'Category', 'Preconditions',
                'Test_Steps', 'Test_Data', 'Expected_Result', 'FRD_Reference', 'Generated_On'
            ]
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                if self.all_test_cases:
                    for test_case in self.all_test_cases:
                        # Map fields to CSV columns
                        csv_row = {
                            'Test_Case_ID': test_case.get('test_case_id', test_case.get('id', '')),
                            'Test_Case_Name': test_case.get('test_case_name', test_case.get('name', '')),
                            'Feature_ID': test_case.get('feature_id', ''),
                            'Feature_Name': test_case.get('feature_name', ''),
                            'Module': test_case.get('module', ''),
                            'Test_Type': test_case.get('test_type', test_case.get('type', '')),
                            'Priority': test_case.get('priority', ''),
                            'Category': test_case.get('category', ''),
                            'Preconditions': test_case.get('preconditions', ''),
                            'Test_Steps': test_case.get('test_steps_formatted', test_case.get('test_steps', '')),
                            'Test_Data': test_case.get('test_data', ''),
                            'Expected_Result': test_case.get('expected_result', ''),
                            'FRD_Reference': test_case.get('frd_reference', ''),
                            'Generated_On': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        writer.writerow(csv_row)
                else:
                    # Write empty row if no test cases
                    empty_row = {field: '' for field in fieldnames}
                    writer.writerow(empty_row)
            
            logger.info(f"CSV file saved: {filepath}")
            print(f"✅ CSV saved successfully at: {filepath}")
            return filepath, f"CSV file saved successfully: {filename}"
            
        except Exception as e:
            logger.error(f"Error saving CSV: {str(e)}")
            print(f"❌ Error saving CSV: {str(e)}")
            return None, f"Error saving CSV: {str(e)}"
    
    def get_summary_stats(self):
        """Get summary statistics using native Python"""
        if not self.all_test_cases:
            return {
                'total_test_cases': 0,
                'test_types': {},
                'priorities': {},
                'categories': {},
                'modules': {},
                'features': {}
            }
        
        stats = {
            'total_test_cases': len(self.all_test_cases),
            'test_types': {},
            'priorities': {},
            'categories': {},
            'modules': {},
            'features': {}
        }
        
        # Count occurrences using native Python
        for tc in self.all_test_cases:
            # Test types
            test_type = tc.get('test_type', tc.get('type', 'Unknown'))
            stats['test_types'][test_type] = stats['test_types'].get(test_type, 0) + 1
            
            # Priorities  
            priority = tc.get('priority', 'Unknown')
            stats['priorities'][priority] = stats['priorities'].get(priority, 0) + 1
            
            # Categories
            category = tc.get('category', 'Unknown')
            stats['categories'][category] = stats['categories'].get(category, 0) + 1
            
            # Modules
            module = tc.get('module', 'Unknown')
            stats['modules'][module] = stats['modules'].get(module, 0) + 1
            
            # Features
            feature = tc.get('feature_name', 'Unknown')
            stats['features'][feature] = stats['features'].get(feature, 0) + 1
        
        return stats
    
    def export_to_csv(self, download_folder):
        """Export test cases to CSV using native Python"""
        if not self.all_test_cases:
            return None, "No test cases to export"
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"generated_testcases_{timestamp}.csv"
        filepath = os.path.join(download_folder, filename)
        
        try:
            # Define standard field order for better readability
            standard_fields = [
                'feature_name', 'feature_id', 'module', 'test_case_id', 'test_case_name',
                'description', 'priority', 'test_type', 'complexity', 'prerequisites',
                'test_steps', 'expected_result', 'test_data', 'notes'
            ]
            
            # Get all unique field names from data
            all_fields = set()
            for test_case in self.all_test_cases:
                all_fields.update(test_case.keys())
            
            # Combine standard fields with any additional fields found
            fieldnames = []
            for field in standard_fields:
                if field in all_fields:
                    fieldnames.append(field)
                    all_fields.remove(field)
            
            # Add any remaining fields
            fieldnames.extend(sorted(all_fields))
            
            # Write CSV with proper formatting
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                # Write data with formatted list fields
                for test_case in self.all_test_cases:
                    formatted_case = {}
                    for field in fieldnames:
                        value = test_case.get(field, '')
                        
                        # Format list fields properly
                        if field in ['test_steps', 'prerequisites'] and isinstance(value, list):
                            if field == 'test_steps':
                                value = '; '.join([f"{i+1}. {step}" for i, step in enumerate(value)])
                            else:
                                value = '; '.join(value)
                        
                        formatted_case[field] = str(value) if value else ''
                    
                    writer.writerow(formatted_case)
            
            return filename, "Test cases exported successfully"
        except Exception as e:
            logger.error(f"Error exporting CSV: {e}")
            return None, f"Error exporting CSV: {str(e)}"
    
    def export_to_excel(self, download_folder):
        """Export test cases to Excel using openpyxl"""
        if not self.all_test_cases:
            return None, "No test cases to export"
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"generated_testcases_{timestamp}.xlsx"
        filepath = os.path.join(download_folder, filename)
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Cases"
            
            # Define standard headers for better structure
            standard_headers = [
                'feature_name', 'feature_id', 'module', 'test_case_id', 'test_case_name',
                'description', 'priority', 'test_type', 'complexity', 'prerequisites',
                'test_steps', 'expected_result', 'test_data', 'notes'
            ]
            
            # Create display headers
            display_headers = [
                'Feature Name', 'Feature ID', 'Module', 'Test Case ID', 'Test Case Name',
                'Description', 'Priority', 'Test Type', 'Complexity', 'Prerequisites',
                'Test Steps', 'Expected Result', 'Test Data', 'Notes'
            ]
            
            # Add formatted headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col, header in enumerate(display_headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data with proper formatting
            for row_idx, test_case in enumerate(self.all_test_cases, 2):
                for col_idx, field in enumerate(standard_headers, 1):
                    value = test_case.get(field, '')
                    
                    # Format list fields properly
                    if field in ['test_steps', 'prerequisites'] and isinstance(value, list):
                        if field == 'test_steps':
                            value = '\n'.join([f"{i+1}. {step}" for i, step in enumerate(value)])
                        else:
                            value = '\n'.join(value)
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add statistics sheet
            stats_ws = wb.create_sheet("Statistics")
            stats = self.get_statistics()
            
            # Format statistics sheet
            stats_ws.cell(row=1, column=1, value="Metric").font = Font(bold=True)
            stats_ws.cell(row=1, column=2, value="Value").font = Font(bold=True)
            
            for row_idx, (metric, value) in enumerate(stats.items(), 2):
                stats_ws.cell(row=row_idx, column=1, value=metric)
                stats_ws.cell(row=row_idx, column=2, value=str(value))
            
            wb.save(filepath)
            
            return filename, "Test cases exported to Excel successfully"
        except ImportError:
            logger.warning("openpyxl not available, falling back to CSV format")
            return self.export_to_csv_fallback(filepath.replace('.xlsx', '.csv'))
        except Exception as e:
            logger.error(f"Error exporting Excel: {e}")
            return None, f"Error exporting Excel: {str(e)}"
    
    def export_to_csv_fallback(self, filepath):
        """Fallback CSV export method"""
        try:
            filename = os.path.basename(filepath)
            
            # Define standard field order
            standard_fields = [
                'feature_name', 'feature_id', 'module', 'test_case_id', 'test_case_name',
                'description', 'priority', 'test_type', 'complexity', 'prerequisites',
                'test_steps', 'expected_result', 'test_data', 'notes'
            ]
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=standard_fields, extrasaction='ignore')
                writer.writeheader()
                
                for test_case in self.all_test_cases:
                    # Format data for CSV
                    formatted_case = {}
                    for field in standard_fields:
                        value = test_case.get(field, '')
                        
                        # Format list fields properly
                        if isinstance(value, list):
                            if field == 'test_steps':
                                value = '; '.join([f"{i+1}. {step}" for i, step in enumerate(value)])
                            else:
                                value = '; '.join(value)
                        
                        formatted_case[field] = str(value) if value else ''
                    
                    writer.writerow(formatted_case)
            
            return filename, "Test cases exported to CSV successfully (fallback mode)"
        except Exception as e:
            logger.error(f"Error in CSV fallback export: {e}")
            return None, f"Error in CSV fallback export: {str(e)}"
